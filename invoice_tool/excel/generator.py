"""Layer 5 — Excel Report Generator.

Uses template.xlsx as base, dynamically expands engineer blocks,
preserves formatting, and writes all computed values.
Excel formulas are NOT relied upon — all values are pre-computed in Python.
"""

from __future__ import annotations

import copy
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from invoice_tool.models import (
    Category,
    EngineerBlock,
    InvoiceResult,
    StrictValidationError,
)

# Template layout constants (based on analysis of Calculation 1535984.xlsx)
TITLE_ROW = 1
HEADER_ROW_1 = 2  # Project Name, Emerson Project #
HEADER_ROW_2 = 3  # Purchase Order #, Invoice #
ENGINEER_HEADER_ROW = 4  # Engineer names (merged across 3 cols each)
SUB_HEADER_ROW = 5  # Normal, Overtime, HOT
DATA_START_ROW = 6
DATA_END_ROW = 27  # rows 6-27 for dates (22 date slots)
TOTAL_HOURS_ROW = 28
HOURLY_RATE_ROW = 29
BILL_AMOUNT_ROW = 30
TOTAL_INVOICE_ROW = 31

# Summary sections
INVOICE_DESC_ROW = 33
NORMAL_DETAIL_HEADER_ROW = 34
NORMAL_DETAIL_COL_HEADER_ROW = 35
NORMAL_DETAIL_START_ROW = 36

# Formatting constants
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

HEADER_FONT = Font(name='Calibri', size=11, bold=True)
DATA_FONT = Font(name='Calibri', size=11)
TITLE_FONT = Font(name='Calibri', size=12, bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
DOLLAR_FORMAT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
NUMBER_FORMAT = '#,##0.00'
DATE_FORMAT = '[$-F800]dddd, mmmm dd, yyyy'


def _copy_cell_style(source_cell, target_cell):
    """Copy formatting from source cell to target cell."""
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


def generate_excel_report(
    result: InvoiceResult,
    template_path: str | Path,
    output_path: str | Path,
) -> Path:
    """Generate the final Excel invoice report from computed results.

    Uses the template for formatting reference, but writes a fresh workbook
    with dynamically sized engineer blocks.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    # Load template for style reference
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # First unmerge ALL merged cells (we'll re-merge as needed)
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # Now clear all cells from row 4 onward (keep title and project info)
    for row in range(ENGINEER_HEADER_ROW, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None

    # Clear engineer header row sub-columns
    for col in range(2, ws.max_column + 1):
        ws.cell(row=ENGINEER_HEADER_ROW, column=col).value = None
        ws.cell(row=SUB_HEADER_ROW, column=col).value = None

    blocks = result.engineer_blocks
    num_engineers = len(blocks)

    # Each engineer gets 3 columns: Normal, OT, HOT
    # Starting from column B (col 2)
    total_data_cols = num_engineers * 3
    last_data_col = 1 + total_data_cols  # col A is dates

    # Re-merge title row to span all columns
    if last_data_col > 1:
        # Unmerge old title
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == 1 and mr.max_row == 1:
                ws.unmerge_cells(str(mr))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_data_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = 'Invoice Summary for Hours worked by Engineers'
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER_ALIGN

    # --- Write header rows ---
    ws.cell(row=HEADER_ROW_1, column=1).value = 'Project Name'
    ws.cell(row=HEADER_ROW_1, column=1).font = HEADER_FONT

    # Unmerge and re-merge B2 area
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == HEADER_ROW_1 and mr.min_column == 2:
            ws.unmerge_cells(str(mr))
    ws.cell(row=HEADER_ROW_1, column=2).value = 'Saipem'
    ws.cell(row=HEADER_ROW_1, column=2).font = HEADER_FONT

    ws.cell(row=HEADER_ROW_2, column=1).value = 'Purchase Order #'
    ws.cell(row=HEADER_ROW_2, column=1).font = HEADER_FONT

    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == HEADER_ROW_2 and mr.min_column == 2:
            ws.unmerge_cells(str(mr))
    ws.cell(row=HEADER_ROW_2, column=2).value = int(result.po_data.contract_number)
    ws.cell(row=HEADER_ROW_2, column=2).font = HEADER_FONT

    # --- Write engineer headers (row 4 & 5) ---
    for i, block in enumerate(blocks):
        start_col = 2 + i * 3  # B, E, H, K, ...

        # Engineer name header (merged across 3 cols)
        suffix = "Offshore" if block.category == Category.OFFSHORE else "Onshore"
        eng_header = f"{block.name}-{suffix}"
        ws.merge_cells(
            start_row=ENGINEER_HEADER_ROW, start_column=start_col,
            end_row=ENGINEER_HEADER_ROW, end_column=start_col + 2,
        )
        cell = ws.cell(row=ENGINEER_HEADER_ROW, column=start_col)
        cell.value = eng_header
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

        # Sub-headers: Normal, Overtime, HOT
        for j, label in enumerate(["Normal ", "Overtime", "HOT"]):
            c = ws.cell(row=SUB_HEADER_ROW, column=start_col + j)
            c.value = label
            c.font = DATA_FONT
            c.alignment = CENTER_ALIGN

    # Date column header
    ws.cell(row=ENGINEER_HEADER_ROW, column=1).value = 'Date'
    ws.cell(row=ENGINEER_HEADER_ROW, column=1).font = HEADER_FONT
    ws.cell(row=ENGINEER_HEADER_ROW, column=1).alignment = CENTER_ALIGN

    # --- Determine date rows needed ---
    all_dates = result.all_dates
    num_date_rows = max(len(all_dates), DATA_END_ROW - DATA_START_ROW + 1)

    # Adjust row references if we need more rows than template
    actual_data_end = DATA_START_ROW + num_date_rows - 1
    total_row = actual_data_end + 1
    rate_row = total_row + 1
    bill_row = rate_row + 1
    total_invoice_row = bill_row + 1

    # --- Write date data ---
    for date_idx, dt in enumerate(all_dates):
        row_num = DATA_START_ROW + date_idx
        date_cell = ws.cell(row=row_num, column=1)
        date_cell.value = datetime(dt.year, dt.month, dt.day)
        date_cell.number_format = DATE_FORMAT
        date_cell.alignment = CENTER_ALIGN
        date_cell.font = DATA_FONT

        for eng_idx, block in enumerate(blocks):
            start_col = 2 + eng_idx * 3
            # Find entry for this date
            entry = None
            for e in block.entries:
                if e.date == dt:
                    entry = e
                    break

            if entry:
                for j, hours in enumerate([entry.normal_hours, entry.ot_hours, entry.hot_hours]):
                    if hours > 0:
                        c = ws.cell(row=row_num, column=start_col + j)
                        c.value = float(hours)
                        c.alignment = CENTER_ALIGN
                        c.font = DATA_FONT

    # --- Write Total Hours row ---
    ws.cell(row=total_row, column=1).value = 'Total Hours '
    ws.cell(row=total_row, column=1).font = HEADER_FONT

    for eng_idx, block in enumerate(blocks):
        start_col = 2 + eng_idx * 3
        for j, total in enumerate([block.total_normal_hours, block.total_ot_hours, block.total_hot_hours]):
            c = ws.cell(row=total_row, column=start_col + j)
            c.value = float(total)
            c.font = HEADER_FONT
            c.alignment = CENTER_ALIGN
            c.number_format = NUMBER_FORMAT

    # --- Write Hourly Rate row ---
    ws.cell(row=rate_row, column=1).value = 'Hourly Rate '
    ws.cell(row=rate_row, column=1).font = HEADER_FONT

    for eng_idx, block in enumerate(blocks):
        start_col = 2 + eng_idx * 3
        for j, rate in enumerate([block.normal_rate, block.ot_rate, block.hot_rate]):
            c = ws.cell(row=rate_row, column=start_col + j)
            c.value = float(rate)
            c.font = DATA_FONT
            c.alignment = CENTER_ALIGN
            c.number_format = NUMBER_FORMAT

    # --- Write Bill Amount row ---
    ws.cell(row=bill_row, column=1).value = ' Bill Amount'
    ws.cell(row=bill_row, column=1).font = HEADER_FONT

    for eng_idx, block in enumerate(blocks):
        start_col = 2 + eng_idx * 3
        costs = [block.normal_cost, block.ot_cost, block.hot_cost]
        for j, cost in enumerate(costs):
            c = ws.cell(row=bill_row, column=start_col + j)
            c.value = float(cost)
            c.font = DATA_FONT
            c.alignment = CENTER_ALIGN
            c.number_format = NUMBER_FORMAT

    # --- Write Total Invoice Amount row ---
    ws.cell(row=total_invoice_row, column=1).value = 'Total Invoice Amount'
    ws.cell(row=total_invoice_row, column=1).font = HEADER_FONT

    # Merge across all data columns for total
    ws.merge_cells(
        start_row=total_invoice_row, start_column=2,
        end_row=total_invoice_row, end_column=last_data_col,
    )
    total_cell = ws.cell(row=total_invoice_row, column=2)
    total_cell.value = float(result.grand_total)
    total_cell.font = HEADER_FONT
    total_cell.alignment = CENTER_ALIGN
    total_cell.number_format = NUMBER_FORMAT

    # --- INVOICE DESCRIPTION section ---
    desc_start = total_invoice_row + 2
    _write_summary_section(ws, desc_start, blocks, result, last_data_col)

    # --- Set column widths ---
    ws.column_dimensions['A'].width = 33
    for col in range(2, last_data_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Save
    wb.save(str(output_path))
    return output_path


def _write_summary_section(
    ws,
    start_row: int,
    blocks: list[EngineerBlock],
    result: InvoiceResult,
    last_col: int,
) -> None:
    """Write the invoice description/summary section below the main grid."""
    row = start_row
    # Section header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1).value = 'INVOICE DESCRIPTION'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    # Rate Chart header
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    ws.cell(row=row, column=8).value = 'Rate Chart'
    ws.cell(row=row, column=8).font = HEADER_FONT

    row += 1
    # Normal Hours section
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1).value = 'Normal Hours / Per Day Calculation'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    ws.cell(row=row, column=8).value = 'Description'
    ws.cell(row=row, column=8).font = HEADER_FONT
    ws.cell(row=row, column=9).value = 'Amount (in US $)'
    ws.cell(row=row, column=9).font = HEADER_FONT

    row += 1
    # Column headers
    for col, label in [(1, 'Name of Engineer'), (2, 'Hours Worked'), (3, 'Hourly Rate'), (5, 'Total')]:
        ws.cell(row=row, column=col).value = label
        ws.cell(row=row, column=col).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = CENTER_ALIGN

    row += 1
    # Normal hours per engineer
    for block in blocks:
        suffix = "Offshore" if block.category == Category.OFFSHORE else "Onshore"
        ws.cell(row=row, column=1).value = f"{block.name}-{suffix}"
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).value = float(block.total_normal_hours)
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).value = float(block.normal_rate)
        ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=5).value = float(block.normal_cost)
        ws.cell(row=row, column=5).number_format = DOLLAR_FORMAT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN
        row += 1

    # Total (A)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = 'Total (A)'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    total_a = sum((b.normal_cost for b in blocks), Decimal("0"))
    ws.cell(row=row, column=5).value = float(total_a)
    ws.cell(row=row, column=5).font = HEADER_FONT
    ws.cell(row=row, column=5).number_format = DOLLAR_FORMAT
    ws.cell(row=row, column=5).alignment = CENTER_ALIGN

    row += 2
    # OT section
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1).value = 'OT details'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    row += 1
    for col, label in [(1, 'Name of Engineer'), (2, 'OT Hours'), (3, 'OT rate'), (5, 'Total')]:
        ws.cell(row=row, column=col).value = label
        ws.cell(row=row, column=col).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = CENTER_ALIGN

    # Summary stats
    ws.cell(row=row, column=8).value = 'Total Hours:'
    ws.cell(row=row, column=8).font = HEADER_FONT
    ws.cell(row=row, column=9).value = float(result.total_hours)
    ws.cell(row=row, column=9).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=9).alignment = CENTER_ALIGN

    row += 1
    for block in blocks:
        suffix = "Offshore" if block.category == Category.OFFSHORE else "Onshore"
        ws.cell(row=row, column=1).value = f"{block.name}-{suffix}"
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).value = float(block.total_ot_hours)
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).value = float(block.ot_rate)
        ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=5).value = float(block.ot_cost)
        ws.cell(row=row, column=5).number_format = DOLLAR_FORMAT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN

        # Summary hours on right side
        if block == blocks[0]:
            ws.cell(row=row, column=8).value = 'Normal Hours:'
            ws.cell(row=row, column=9).value = float(result.total_normal_hours)
            ws.cell(row=row, column=9).number_format = NUMBER_FORMAT
            ws.cell(row=row, column=9).alignment = CENTER_ALIGN
        elif block == blocks[1] if len(blocks) > 1 else False:
            ws.cell(row=row, column=8).value = 'OT:'
            ws.cell(row=row, column=9).value = float(result.total_ot_hours)
            ws.cell(row=row, column=9).number_format = NUMBER_FORMAT
            ws.cell(row=row, column=9).alignment = CENTER_ALIGN
        elif block == blocks[2] if len(blocks) > 2 else False:
            ws.cell(row=row, column=8).value = 'HOT:'
            ws.cell(row=row, column=9).value = float(result.total_hot_hours)
            ws.cell(row=row, column=9).number_format = NUMBER_FORMAT
            ws.cell(row=row, column=9).alignment = CENTER_ALIGN

        row += 1

    # Total (B)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = 'Total (B)'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    total_b = sum((b.ot_cost for b in blocks), Decimal("0"))
    ws.cell(row=row, column=5).value = float(total_b)
    ws.cell(row=row, column=5).font = HEADER_FONT
    ws.cell(row=row, column=5).number_format = DOLLAR_FORMAT
    ws.cell(row=row, column=5).alignment = CENTER_ALIGN

    row += 2
    # HOT section
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1).value = 'HOT details'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN

    row += 1
    for col, label in [(1, 'Name of Engineer'), (2, 'HOT Hours'), (3, 'HOT rate'), (5, 'Total')]:
        ws.cell(row=row, column=col).value = label
        ws.cell(row=row, column=col).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = CENTER_ALIGN

    row += 1
    for block in blocks:
        suffix = "Offshore" if block.category == Category.OFFSHORE else "Onshore"
        ws.cell(row=row, column=1).value = f"{block.name}-{suffix}"
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).value = float(block.total_hot_hours)
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).value = float(block.hot_rate)
        ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=5).value = float(block.hot_cost)
        ws.cell(row=row, column=5).number_format = DOLLAR_FORMAT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN
        row += 1

    # Total (C)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1).value = 'Total (C)'
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    total_c = sum((b.hot_cost for b in blocks), Decimal("0"))
    ws.cell(row=row, column=5).value = float(total_c)
    ws.cell(row=row, column=5).font = HEADER_FONT
    ws.cell(row=row, column=5).number_format = DOLLAR_FORMAT
    ws.cell(row=row, column=5).alignment = CENTER_ALIGN

    row += 2
    # TOTAL INVOICE VALUE
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=3)
    cell = ws.cell(row=row, column=1)
    cell.value = 'TOTAL INVOICE VALUE (A+B+C)'
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN

    ws.merge_cells(start_row=row, start_column=5, end_row=row + 1, end_column=5)
    total_cell = ws.cell(row=row, column=5)
    total_cell.value = float(result.grand_total)
    total_cell.font = HEADER_FONT
    total_cell.number_format = DOLLAR_FORMAT
    total_cell.alignment = CENTER_ALIGN
