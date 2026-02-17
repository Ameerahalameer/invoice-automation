"""Tests for Excel report generator."""

import pytest
import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

from invoice_tool.models import (
    Category,
    EngineerBlock,
    EngineerLevel,
    InvoiceResult,
    POData,
    RateSet,
    TimesheetEntry,
)
from invoice_tool.excel.generator import (
    generate_excel_report,
    DATA_START_ROW,
    DATA_END_ROW,
)

TEMPLATE_PATH = Path(__file__).parent.parent / "Calculation 1535984.xlsx"

# Match the generator's row calculation logic
_MIN_DATE_ROWS = DATA_END_ROW - DATA_START_ROW + 1  # 22


def _make_po() -> POData:
    return POData(
        contract_number="1535984",
        onshore_rates={
            EngineerLevel.SERVICE_FIELD: RateSet(
                normal=Decimal("286"), ot=Decimal("372"), hot=Decimal("443")
            ),
        },
        offshore_rates={
            EngineerLevel.SERVICE_FIELD: RateSet(
                normal=Decimal("372"), ot=Decimal("484"), hot=Decimal("577")
            ),
        },
        onshore_hours_per_day=10,
        offshore_hours_per_day=12,
        max_amount_usd=Decimal("999999"),
        source_file="test.pdf",
    )


def _make_entry(name, dt, n, ot, hot, cat):
    return TimesheetEntry(
        engineer_name=name,
        date=dt,
        normal_hours=Decimal(str(n)),
        ot_hours=Decimal(str(ot)),
        hot_hours=Decimal(str(hot)),
        category=cat,
        engineer_level=EngineerLevel.SERVICE_FIELD,
        source_file="test.pdf",
    )


def _make_result_single_engineer() -> InvoiceResult:
    entries = [
        _make_entry("Alice", date(2026, 1, 11), 10, 2, 0, Category.ONSHORE),
        _make_entry("Alice", date(2026, 1, 12), 10, 0, 0, Category.ONSHORE),
    ]
    block = EngineerBlock(
        name="Alice",
        category=Category.ONSHORE,
        engineer_level=EngineerLevel.SERVICE_FIELD,
        entries=entries,
        normal_rate=Decimal("286"),
        ot_rate=Decimal("372"),
        hot_rate=Decimal("443"),
    )
    return InvoiceResult(
        po_data=_make_po(),
        engineer_blocks=[block],
        all_dates=[date(2026, 1, 11), date(2026, 1, 12)],
    )


def _make_result_two_engineers() -> InvoiceResult:
    entries_a = [
        _make_entry("Alice", date(2026, 1, 11), 10, 2, 0, Category.ONSHORE),
    ]
    entries_b = [
        _make_entry("Bob", date(2026, 1, 11), 12, 0, 0, Category.OFFSHORE),
        _make_entry("Bob", date(2026, 1, 16), 0, 0, 8, Category.OFFSHORE),  # Friday
    ]
    block_a = EngineerBlock(
        name="Alice",
        category=Category.ONSHORE,
        engineer_level=EngineerLevel.SERVICE_FIELD,
        entries=entries_a,
        normal_rate=Decimal("286"),
        ot_rate=Decimal("372"),
        hot_rate=Decimal("443"),
    )
    block_b = EngineerBlock(
        name="Bob",
        category=Category.OFFSHORE,
        engineer_level=EngineerLevel.SERVICE_FIELD,
        entries=entries_b,
        normal_rate=Decimal("372"),
        ot_rate=Decimal("484"),
        hot_rate=Decimal("577"),
    )
    return InvoiceResult(
        po_data=_make_po(),
        engineer_blocks=[block_a, block_b],
        all_dates=[date(2026, 1, 11), date(2026, 1, 16)],
    )


@pytest.fixture
def tmp_output(tmp_path):
    return tmp_path / "test_output.xlsx"


class TestExcelGenerator:
    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_output_file_created(self, tmp_output):
        result = _make_result_single_engineer()
        out = generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        assert out.exists()

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_single_engineer_total_hours(self, tmp_output):
        result = _make_result_single_engineer()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        # num_date_rows = max(2, 22) = 22; total_row = 6 + 22 - 1 + 1 = 28
        num_date_rows = max(len(result.all_dates), _MIN_DATE_ROWS)
        total_row = DATA_START_ROW + num_date_rows  # 28
        assert ws.cell(row=total_row, column=2).value == 20.0  # Normal
        assert ws.cell(row=total_row, column=3).value == 2.0   # OT
        assert ws.cell(row=total_row, column=4).value == 0.0   # HOT

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_single_engineer_rates(self, tmp_output):
        result = _make_result_single_engineer()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        num_date_rows = max(len(result.all_dates), _MIN_DATE_ROWS)
        rate_row = DATA_START_ROW + num_date_rows + 1
        assert ws.cell(row=rate_row, column=2).value == 286.0
        assert ws.cell(row=rate_row, column=3).value == 372.0
        assert ws.cell(row=rate_row, column=4).value == 443.0

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_single_engineer_bill_amount(self, tmp_output):
        result = _make_result_single_engineer()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        num_date_rows = max(len(result.all_dates), _MIN_DATE_ROWS)
        bill_row = DATA_START_ROW + num_date_rows + 2
        assert ws.cell(row=bill_row, column=2).value == 5720.0   # 20 * 286
        assert ws.cell(row=bill_row, column=3).value == 744.0    # 2 * 372
        assert ws.cell(row=bill_row, column=4).value == 0.0      # 0 * 443

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_grand_total(self, tmp_output):
        result = _make_result_single_engineer()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        num_date_rows = max(len(result.all_dates), _MIN_DATE_ROWS)
        total_invoice_row = DATA_START_ROW + num_date_rows + 3
        assert ws.cell(row=total_invoice_row, column=2).value == 6464.0  # 5720 + 744

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_two_engineers_column_layout(self, tmp_output):
        result = _make_result_two_engineers()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        # Engineer 1 (Alice) in cols B,C,D (2,3,4)
        # Engineer 2 (Bob) in cols E,F,G (5,6,7)
        # Sub-header row 5
        assert ws.cell(row=5, column=2).value == "Normal "
        assert ws.cell(row=5, column=3).value == "Overtime"
        assert ws.cell(row=5, column=4).value == "HOT"
        assert ws.cell(row=5, column=5).value == "Normal "
        assert ws.cell(row=5, column=6).value == "Overtime"
        assert ws.cell(row=5, column=7).value == "HOT"

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_two_engineers_grand_total(self, tmp_output):
        result = _make_result_two_engineers()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        # Grand total = Alice(10*286 + 2*372) + Bob(12*372 + 8*577)
        # = (2860 + 744) + (4464 + 4616) = 3604 + 9080 = 12684
        num_date_rows = max(len(result.all_dates), _MIN_DATE_ROWS)
        total_invoice_row = DATA_START_ROW + num_date_rows + 3
        assert ws.cell(row=total_invoice_row, column=2).value == 12684.0

    @pytest.mark.skipif(not TEMPLATE_PATH.exists(), reason="Template not available")
    def test_contract_number_written(self, tmp_output):
        result = _make_result_single_engineer()
        generate_excel_report(result, TEMPLATE_PATH, tmp_output)
        wb = openpyxl.load_workbook(str(tmp_output))
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 1535984
